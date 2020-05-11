from django.shortcuts import render
from django.shortcuts import render
from django.http import HttpResponse
from django.http import JsonResponse
from rest_framework.views import APIView,View
from rest_framework.response import Response
# Create your views here.
from django.core.exceptions import MultipleObjectsReturned, ObjectDoesNotExist
import json
from django.core.serializers import serialize
from django.template import loader
from django.http import HttpResponseRedirect
from django.urls import reverse
from django.views import generic
import pandas as pd
from django.http import HttpResponse, HttpResponseRedirect, HttpResponseBadRequest
from django.shortcuts import get_object_or_404, render
from django.urls import reverse
from django.utils import timezone
from .models import Tracker
from django.views.decorators.csrf import csrf_exempt, csrf_protect
#from .models import ModelWithFileField
from pyexcel_xlsx import get_data as xlsx_get
from django.utils.datastructures import MultiValueDictKeyError
from django.core.exceptions import MultipleObjectsReturned
import openpyxl
from django.shortcuts import redirect
import datetime
import pandas as pd
import uuid
from django.contrib.auth import authenticate
from django.contrib.auth import login as auth_login
from django.contrib.auth import logout as auth_logout
from django.views.generic import TemplateView
from django.core import serializers
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import permission_required
import xlwt

class HomeView(View):
    def get(self,request,*args,**kwargs):
        return render(request,'test_cases/login.html')

class DashboardView(View):
    def get(self,request,*args,**kwargs):
        return render(request,'test_cases/dashboard.html')

class ChartData(APIView):
    authentication_classes = []
    permission_classes = []

    def get(self, request, format=None):
        """
        Return a list of all users.
        """
        labels =  self.get_modules()
        data =  self.get_module_count(labels)
        context = {
            'labels':labels,
            'data':data
        }
        return Response(context)

    def get_modules(self):
        records  = Tracker.objects.all()
        mod = set()
        for i in records:
            mod.add(i.module)
        return list(mod)

    def get_module_count(self,mod):
        data = []
        for i in mod:
            data.append(Tracker.objects.filter(module=i).count())
        return data

class LoginView(View):

    def get(self,request,*args,**kwargs):
        return render(request,'test_cases/login.html')
    
    def post(self,request,*args,**kwargs):
        username = request.POST['name']
        password = request.POST['password']
        #request.session['name'] 
        user = authenticate(request, username=username, password=password)
        if user is not None:
            auth_login(request,user)
            return HttpResponseRedirect('/track/browse/')
        else:
            return render(request,'test_cases/login.html',{'message':'Invalid Username or Password'})

class LogoutView(View):

    def get(self,request,*args,**kwargs):
        auth_logout(request)
        return render(request,'test_cases/logout.html')

class UploadView(View):
    def get(self,request,*args,**kwargs):
        return render(request, 'test_cases/index.html')
    
    def post(self,request,*args,**kwargs):
        self.load_workbook(request)
        return HttpResponseRedirect("/track")

    def load_workbook(self,request):
        try:
            excel_file = request.FILES["file"]
            wb = openpyxl.load_workbook(excel_file)
            sheet = pd.read_excel(excel_file)
            worksheet = wb["Sheet1"]
            excel_data = list()
            import_id = str(uuid.uuid1())
            for i in sheet.iterrows():
                row = pd.Series(i[1])
                request.session['uuid'] = import_id
                request.session['name'] = request.user.get_full_name()
                request.session['category'] = request.POST['category']
                case = row['Scenario'] #row_data[0]
                case_id = row['Kainos_ID']#row_data[1]
                name = request.user.get_full_name()
                module = request.POST['module']
                email = request.user.email
                category = request.POST['category']
                try:
                    obj = Tracker.objects.get(module=module,kainos_id=case_id,scenario=case)
                    obj = Tracker.objects.filter(module=module,kainos_id=case_id,scenario=case).update(name=name,email=email,uuid= import_id,category=category)
                except(ObjectDoesNotExist):
                    obj = Tracker(name=name, email=email, module=module,kainos_id=case_id, scenario=case,uuid=import_id,category=category)
                    obj.save()
        
        except(NameError):
            return render(request, 'test_cases/error.html', {'error': 'Could not load the sheet'})



@permission_required('test_cases.can_view', login_url='/login/')
def track(request):
    try:
        if("GET" == request.method):
            import_id = request.session['uuid']
            category = request.session['category']
            obj = Tracker.objects.filter(uuid=import_id,category=category)
            #choices = Tracker._meta.get_field('result').choices
            choices = ['Select','Pass','Fail']
            return render(request, 'test_cases/track.html', {'sheet': obj,'choices':choices})
        elif("POST"==request.method):
            import_id = request.session['uuid']
            obj = Tracker.objects.filter(uuid = import_id)
            kainos = 'kainos_id'
            res = 'result'
            records = len(obj)+1
            l = []
            for i in range(1,records):
                kainos_id = request.POST[kainos+str(i)]
                result = request.POST[res+str(i)]
                record = obj.get(kainos_id=kainos_id)
                record.result=result
                record.last_modified = timezone.now()
                record.save()
                l.append(record)
            return redirect('browse/')
                #record = obj.get(kainos_id=kainos_id)
                #record.result = result
                #record.last_modified = datetime.now()
                #record.save()
                #return render(request, 'test_cases/error.html', {'error': record})

    except(NameError):
            return render(request, 'test_cases/error.html', {'error': 'Could not load the sheet'})

    
    #elif("POST" == request.method):
        #return render(request, 'test_cases/error.html', {'error':"result"})

        #kainos_id = request.POST['kainos_id1']
        #results=Tracker._meta.get_field('result').choices
        
        #return render(request, 'test_cases/error.html', {'error': kainos_id})
        #obj = Tracker(kainos_id=kainos_id,result=result,last_modified=datetime.now)
        
        #obj.last_modified = datetime.now()
        #obj = Tracker.objects.all()
@permission_required('test_cases.can_view', login_url='/login/')
def browse(request):
    if(request.method=='GET'):
        records  = Tracker.objects.all()
        mod = set()
        for i in records:
            mod.add(i.module)
        dicts = get_files_module(mod)

        return render(request, 'test_cases/browse.html', {'module': mod,'dict':dicts,'HCM':['Anunay','OTP']})

def get_files_module(mod):
    '''data = {}
    for i in mod:
        categories = Tracker.objects.filter(module=i)
        cat_list = set()
        for cat in categories:
            cat_list.add(cat.category)
        data[i]=list(cat_list)
    return data
    '''
    data = {}
    for i in mod:
        cat1 = {}
        categories = Tracker.objects.filter(module=i)
        cat_list = set()
        for cat in categories:
            cat_list.add(cat.category)
        for k,p in zip(cat_list,range(1,len(cat_list)+1)):
            cat1['category'+str(p)] = k    
        data[i]=cat1
    return data

def error(request):
    pass1,fail,count = 0,0,0
    if(request.method=="GET"):
        records = Tracker.objects.filter(module="PCI")
        for i in records:
            count+=1
            if(i.result == 'Pass'):
                pass1+=1
            else:
                fail+=1
        percent = (pass1/count)*100
        return render(request, 'test_cases/error.html', {'module':records,'pass':pass1,'fail':fail,'count':count,'percent':percent})

@permission_required('test_cases.can_view', login_url='/login/')
def edit_category(request,module,category):
    if(request.method=="GET"):
        if(module == 'All'):
            records = Tracker.objects.all()
        else:
            records = Tracker.objects.filter(module=module,category=category)
        return render(request, 'test_cases/edit.html', {'records':records,"module":module,'category':category})
    elif(request.method=="POST"):
        obj = Tracker.objects.filter(module=module,category=category)
        kainos = 'kainos_id'
        res = 'result'
        records = len(obj)+1
        l = []    
        for i in range(1,records):        
            kainos_id = request.POST[kainos+str(i)]
            result = request.POST[res+str(i)]
            record = obj.get(kainos_id=kainos_id)
            record.result=result
            record.last_modified = timezone.now()
            record.save()
            remove_dups()
            l.append(record)
        records  = Tracker.objects.all()
        mod = set()
        for i in records:
            mod.add(i.module)
        return render(request, 'test_cases/browse.html', {'module': mod}) #render(request,'test_cases/module.html',{'records':obj,'module':obj})
  

        

@permission_required('test_cases.can_view', login_url='/login/')
def edit(request,module):
    if(request.method=="GET"):
        if(module == 'All'):
            records = Tracker.objects.all()
        else:
            records = Tracker.objects.filter(module=module)
        return render(request, 'test_cases/edit.html', {'records':records,"module":module})
        #return render(request,'test_cases/error.html',{'name':'Hi!'})
    elif(request.method=="POST"):
        if(module=='All'):
            obj = Tracker.objects.all()
        else:
            obj = Tracker.objects.filter(module=module)
        kainos = 'kainos_id'
        res = 'result'
        records = len(obj)+1
        l = []    
        for i in range(1,records):        
            kainos_id = request.POST[kainos+str(i)]
            result = request.POST[res+str(i)]
            record = obj.get(kainos_id=kainos_id)
            record.result=result
            record.last_modified = timezone.now()
            record.save()
            l.append(record)
        return redirect('/track/browse/') 

def remove_dups():
    for row in Tracker.objects.all():
        if(Tracker.objects.filter(kainos_id=row.kainos_id).count() > 1):
            row.delete()


@permission_required('test_cases.can_view', login_url='/login/')
def category(request,module,category):
    if(request.method=="GET"):
        records = Tracker.objects.filter(module=module,category=category)
        pass1,fail,count,percent = 0,0,0,0
        for i in records:
            count+=1
            if(i.result == 'Pass'):
                pass1+=1
            else:
                fail+=1
        if(count>0):
            percent = (pass1/count)*100
        return render(request,'test_cases/category.html',{'records':records,'module':module,'category':category,'pass':pass1,'fail':fail,'count':count,'percent':percent})
       

@permission_required('test_cases.can_view', login_url='/login/')
def module(request,module):
    if(request.method=="GET"):
        if(module == 'All'):
            records = Tracker.objects.all()
        else:
            records = Tracker.objects.filter(module=module)
        
        pass1,fail,count,percent = 0,0,0,0
        for i in records:
            count+=1
            if(i.result == 'Pass'):
                pass1+=1
            else:
                fail+=1
        if(count>0):
            percent = (pass1/count)*100
        return render(request,'test_cases/module.html',{'records':records,'module':module,'pass':pass1,'fail':fail,'count':count,'percent':percent})


def upload(request):
    try:
        if(request.method == "GET"):
            return render(request, 'test_cases/index.html', {})

        elif(request.method == 'POST'):
            load_workbook(request)
            render(request, 'upload.html', {'successful_submit': True})

    except(NameError, MultiValueDictKeyError, Tracker.DoesNotExist):
        return render(request, 'test_cases/error.html', {'error': 'Please check the information and try again@'})
    except(MultipleObjectsReturned):
        return render(request, 'test_cases/error.html', {'error': 'Multiple Objects Returned'})
    else:
        return HttpResponseRedirect("/track")

'''def index(request):
    #session = get_object_or_404(Tracker)
    try:
        if request.method == "POST":
            def choice_func(row):
                q = Tracker.objects.filter(slug=row[0])[0]
                row[0] = q
                return row
            if form.is_valid():
                request.FILES['file'].save_book_to_database(
                models=[Tracker, LoginForm],
                initializers=[None, choice_func],
                mapdicts=[
                    {
                    "Scenario": "scenario",
                    "Kainos Automated TC ID": "kainos_id",
                    "Last Modified": "last_modified",
                    "Name": "name",
                    "Email": "email",
                    "Module":"Module"
                }]
            )
            return redirect('test_cases/track.html')
        else:
            return HttpResponseBadRequest()

    except(KeyError, Tracker.DoesNotExist,MultiValueDictKeyError):
        return render(request,'test_cases/index.html',{'error_message':'Please check the information and try again@'})

  
@csrf_exempt
def upload_file_view(request):
    return _upload_file_view(request)

@csrf_protect
def _upload_file_view(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            instance = ModelWithFileField(file_field=request.FILES['file'])
            instance.save()
            return HttpResponseRedirect('/main/track/')
    else:
        form = UploadFileForm()
    return render(request, 'test_cases/index.html', {'case': form})
def track(request):
    if "GET" == request.method:
        return render(request, 'test_cases/track.html', {})
    else:
        excel_file = request.FILES["excel_file"]

        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)
        sheet = pd.read_excel(excel_file)

        # getting a particular sheet by name out of many sheets
        worksheet = wb["Sheet1"]
        print(worksheet)

        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            row_data = list()
            #for cell in row:
            #    row_data.append(str(cell.value))
            data = row[0]
            excel_data.append(row_data)
        return render(request,'test_cases/error.html',{'error':sheet.iloc[0,'Kainos Automated TC ID']})
        #return render(request, 'test_cases/track.html', {"excel_data":excel_data})
    
'''
# return render(request, 'test_cases/track.html', {'case': 'Hello'})
