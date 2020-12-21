from django.shortcuts import render
from django.http import HttpResponse
from django.http import JsonResponse
from rest_framework.views import APIView,View
from rest_framework.response import Response
# Create your views here.
from django.core.exceptions import MultipleObjectsReturned, ObjectDoesNotExist
import json
from django.core.serializers import serialize
from django.template import loader
from django.http import HttpResponseRedirect
from django.urls import reverse
from django.views import generic
import pandas as pd
from django.http import HttpResponse, HttpResponseRedirect, HttpResponseBadRequest
from django.shortcuts import get_object_or_404, render
from django.urls import reverse
from django.utils import timezone
from .models import Tracker
from django.views.decorators.csrf import csrf_exempt, csrf_protect
#from .models import ModelWithFileField
from pyexcel_xlsx import get_data as xlsx_get
from django.utils.datastructures import MultiValueDictKeyError
from django.core.exceptions import MultipleObjectsReturned
import openpyxl
from django.shortcuts import redirect
import datetime
import pandas as pd
import uuid
from django.contrib.auth import authenticate
from django.contrib.auth import login as auth_login
from django.contrib.auth import logout as auth_logout
from django.views.generic import TemplateView
from django.core import serializers
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import permission_required

class HomeView(View):
    def get(self,request,*args,**kwargs):
        return render(request,'test_cases/login.html')

class DashboardView(View):
    def get(self,request,*args,**kwargs):
        return render(request,'test_cases/dashboard.html')

class ChartData(APIView):
    authentication_classes = []
    permission_classes = []

    def get(self, request, format=None):
        """
        Return a list of all users.
        """
        labels =  self.get_modules()
        data =  self.get_module_count(labels)
        context = {
            'labels':labels,
            'data':data
        }
        return Response(context)

    def get_modules(self):
        records  = Tracker.objects.all()
        mod = set()
        for i in records:
            mod.add(i.module)
        return list(mod)

    def get_module_count(self,mod):
        data = []
        for i in mod:
            data.append(Tracker.objects.filter(module=i).count())
        return data

class LoginView(View):

    def get(self,request,*args,**kwargs):
        return render(request,'test_cases/login.html')
    
    def post(self,request,*args,**kwargs):
        username = request.POST['name']
        password = request.POST['password']
        #request.session['name'] 
        user = authenticate(request, username=username, password=password)

        if user is not None:
            auth_login(request,user)
            return HttpResponseRedirect('/track/browse/')
        else:
            return render(request,'test_cases/login.html',{'message':'Invalid Username or Password'})

class LogoutView(View):

    def get(self,request,*args,**kwargs):
        auth_logout(request)
        return render(request,'test_cases/logout.html')

class UploadView(View):
    def get(self,request,*args,**kwargs):
        return render(request, 'test_cases/index.html')
    
    def post(self,request,*args,**kwargs):
        self.load_workbook(request)
        return HttpResponseRedirect("/track")

    def load_workbook(self,request):
        try:
            excel_file = request.FILES["file"]
            wb = openpyxl.load_workbook(excel_file)
            sheet = pd.read_excel(excel_file)
            worksheet = wb["Test Scripts"]
            excel_data = list()
            import_id = str(uuid.uuid1())
            for i,ind in zip(sheet.iterrows(),range(len(sheet))):
                row = pd.Series(i[1])
                request.session['uuid'] = import_id
                request.session['name'] = request.user.get_full_name()
                request.session['category'] = row['Category']
                case = row['Test Scenario'] #row_data[0]

                name = row['Tester']
                if not name:
                    name = None
                result = row['Results']
                if(result is "Pass" or  result is "P"):
                    result = "Pass"
                elif(result is "Fail" or result is "F"):
                    result = "Fail"
                else:
                    result = "None"
                module = row['Category']
                category = row['Test Scripts']

                #module = request.POST['module']
                email = request.user.email
                #category = request.POST['category']
                case_id = str(module)+"_"+str(category)+"_"+str(ind)
                try:
                    obj = Tracker.objects.get(module=module,case_id=case_id,scenario=case)
                    obj = Tracker.objects.filter(module=module,case_id=case_id,scenario=case).update(name=name,uuid= import_id,category=category)
                except(ObjectDoesNotExist):
                    obj = Tracker(name=name, module=module,case_id=case_id, scenario=case,uuid=import_id,category=category,result=result)
                    obj.save()
        
        except(NameError):
            return render(request, 'test_cases/error.html', {'error': 'Could not load the sheet'})



@permission_required('test_cases.can_view', login_url='/login/')
def track(request):
    try:
        if("GET" == request.method):
            import_id = request.session['uuid']
            obj = Tracker.objects.filter(uuid=import_id)
            choices = ['None','Pass','Fail']
            return render(request, 'test_cases/track.html', {'sheet': obj,'choices':choices})
        elif("POST"==request.method):
            import_id = request.session['uuid']
            obj = Tracker.objects.filter(uuid = import_id)
            case_id = 'kainos_id'
            res = 'result'
            records = len(obj)+1
            l = []
            for i in range(1,records):
                kainos_id = request.POST[case_id+str(i)]
                result = request.POST[res+str(i)]
                record = obj.get(case_id=kainos_id)
                record.result=result
                record.last_modified = timezone.now()
                record.save()
                l.append(record)
            return redirect('browse/')

    except(NameError):
            return render(request, 'test_cases/error.html', {'error': 'Could not load the sheet'})

    
    
@permission_required('test_cases.can_view', login_url='/login/')
def browse(request):
    if(request.method=='GET'):
        records  = Tracker.objects.all()
        mod = set()
        for i in records:
            mod.add(i.module)
        dicts = get_files_module(mod)

        return render(request, 'test_cases/browse.html', {'module': mod,'dict':dicts})

def get_files_module(mod):
    '''data = {}
    for i in mod:
        categories = Tracker.objects.filter(module=i)
        cat_list = set()
        for cat in categories:
            cat_list.add(cat.category)
        data[i]=list(cat_list)
    return data
    '''
    data = {}
    for i in mod:
        cat1 = {}
        categories = Tracker.objects.filter(module=i)
        cat_list = set()
        for cat in categories:
            cat_list.add(cat.category)
        for k,p in zip(cat_list,range(1,len(cat_list)+1)):
            cat1['category'+str(p)] = k    
        data[i]=cat1
    return data

def error(request):
    pass1,fail,count = 0,0,0
    if(request.method=="GET"):
        records = Tracker.objects.filter(module="PCI")
        for i in records:
            count+=1
            if(i.result == 'Pass'):
                pass1+=1
            else:
                fail+=1
        percent = (pass1/count)*100
        return render(request, 'test_cases/error.html', {'module':records,'pass':pass1,'fail':fail,'count':count,'percent':percent})

@permission_required('test_cases.can_view', login_url='/login/')
def edit_category(request,module,category):
    if(request.method=="GET"):
        if(module == 'All'):
            records = Tracker.objects.all()
        else:
            records = Tracker.objects.filter(module=module,category=category)
        return render(request, 'test_cases/edit.html', {'records':records,"module":module,'category':category})
    elif(request.method=="POST"):
        obj = Tracker.objects.filter(module=module,category=category)
        case_id = 'kainos_id'
        res = 'result'
        records = len(obj)+1
        l = []    
        for i in range(1,records):        
            kainos_id = request.POST[case_id+str(i)]
            result = request.POST[res+str(i)]
            record = obj.get(case_id=kainos_id)
            record.result=result
            record.last_modified = timezone.now()
            record.save()
            remove_dups()
            l.append(record)
        records  = Tracker.objects.all()
        mod = set()
        for i in records:
            mod.add(i.module)
        return render(request, 'test_cases/browse.html', {'module': mod}) #render(request,'test_cases/module.html',{'records':obj,'module':obj})
  

        

@permission_required('test_cases.can_view', login_url='/login/')
def edit(request,module):
    if(request.method=="GET"):
        if(module == 'All'):
            records = Tracker.objects.all()
        else:
            records = Tracker.objects.filter(module=module)
        return render(request, 'test_cases/edit.html', {'records':records,"module":module})
    elif(request.method=="POST"):
        if(module=='All'):
            obj = Tracker.objects.all()
        else:
            obj = Tracker.objects.filter(module=module)
        case_id = 'kainos_id'
        res = 'result'
        records = len(obj)+1
        l = []    
        for i in range(1,records):        
            kainos_id = request.POST[case_id+str(i)]
            result = request.POST[res+str(i)]
            record = obj.get(case_id=kainos_id)
            record.result=result
            record.last_modified = timezone.now()
            record.save()
            l.append(record)
        return redirect('/track/browse/') 

def remove_dups():
    for row in Tracker.objects.all():
        if(Tracker.objects.filter(id=row.id).count() > 1):
            row.delete()


@permission_required('test_cases.can_view', login_url='/login/')
def category(request,module,category):
    if(request.method=="GET"):
        records = Tracker.objects.filter(module=module,category=category)
        pass1,fail,count,percent = 0,0,0,0
        for i in records:
            count+=1
            if(i.result == 'Pass'):
                pass1+=1
            else:
                fail+=1
        if(count>0):
            percent = (pass1/count)*100
        return render(request,'test_cases/category.html',{'records':records,'module':module,'category':category,'pass':pass1,'fail':fail,'count':count,'percent':percent})
       

@permission_required('test_cases.can_view', login_url='/login/')
def module(request,module):
    if(request.method=="GET"):
        if(module == 'All'):
            records = Tracker.objects.all()
        else:
            records = Tracker.objects.filter(module=module)
        
        pass1,fail,count,percent = 0,0,0,0
        for i in records:
            count+=1
            if(i.result == 'Pass'):
                pass1+=1
            else:
                fail+=1
        if(count>0):
            percent = (pass1/count)*100
        return render(request,'test_cases/module.html',{'records':records,'module':module,'pass':pass1,'fail':fail,'count':count,'percent':percent})


